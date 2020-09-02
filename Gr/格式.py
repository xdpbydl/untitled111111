import pandas as pd
import openpyxl
import numpy as np
import ubpa.itools.rpa_time as rpa_time
from openpyxl.utils import get_column_letter, column_index_from_string
from win32com.client import Dispatch


# import ubpa.ifile as ifile


# 另存为 因为Html网页格式表格不能读取，统一修改格式
def Save_Format_xls(filename):
    xlApp = Dispatch("Excel.Application")
    xlApp.Visible = False  # 后台运行
    xlApp.DisplayAlerts = False  # 不警报
    xlBook = xlApp.Workbooks.Open(filename)
    xlBook.SaveAs(filename, 56)  # 56   xls    51 xlsx
    xlBook.Close()


def Save_Formats_xls(filename, newfilename):
    xlApp = Dispatch("Excel.Application")
    xlApp.Visible = False  # 后台运行
    xlApp.DisplayAlerts = False  # 不警报
    xlBook = xlApp.Workbooks.Open(filename)
    xlBook.SaveAs(newfilename, 56)  # 56   xls    51 xlsx
    xlBook.Close()


print(rpa_time.get_current_time())

Save_Format_xls(self.Excel_zhuangxiangInfo_Path)

df1 = pd.read_excel(self.Excel_zhuangxiangInfo_Path)
# 删除不需要的列
df1.drop(["装箱单编号", "版本", "使用单位", "井道图号", "提升高度"], axis=1, inplace=True)

# ['工号', '订货单位', '型号', '层站门', '序号', '箱代码', '零件编号', '零件名称', '图表号', '作业单位','规格', '数量', '备注', '源图号作业', '物资编码', '参考图号']
# 帅选主机数据
df_zhuji = df1.loc[df1["箱代码"].str.strip() == "01-1"]

# 帅选对重块数据  and df1["零件名称"].str.strip() != "14#对重块箱"
df_duizhongkuai = df1.loc[(df1["箱代码"].str.strip() == "14-1") | (df1["箱代码"].str.strip() == "14-2")]
df_duizhongkuai = df_duizhongkuai.loc[df_duizhongkuai["零件名称"].str.strip() != "14#对重块箱"]

Save_Format_xls(self.Excel_xiangtou_Path)
df2 = pd.read_excel(self.Excel_xiangtou_Path, header=1)

# 获取层站门空白数据的索引
dfnaIndex = df2[pd.isna(df2["层站门"])].index

for i in range(len(dfnaIndex)):
    # 空白层站门 = 提升高度 + "MM "
    # df2["层站门"][dfnaIndex[i]] = str(int(df2["提升高度"][dfnaIndex[i]])) + "MM "
    df2.loc[dfnaIndex[i], "层站门"] = str(int(df2["提升高度"][dfnaIndex[i]])) + "MM "

# 箱头sheet表
df_xiangtou = df2[["工号", "订货单位", "型号", "层站门", "箱代码", "箱名", "备注"]]
# df_xiangtou["订货单位"] = df_xiangtou["订货单位"].astype("string")

# print(df_xiangtou.tail())

#  [u"工号",u"订货单位",u"型号",u"层站门"]  借助pivot_table 透视表并且以count聚合 reset_index重置索引得到表结构

df_table = pd.pivot_table(df_xiangtou, index=["工号", "订货单位", "型号", "层站门"], aggfunc='count').reset_index()[
    ["工号", "订货单位", "型号", "层站门"]]
df_table["提货方式"] = "送货"
# print(df_table)


# va = []
wb = openpyxl.load_workbook(self.Excel_tongyongmoban, data_only=False)
sheet = wb["工号信息"]
for i in range(len(df_table)):
    va = df_table.iloc[i].values.T.tolist()
    for ix in range(len(va)):
        sheet[get_column_letter(ix + 1) + str(i + 2)] = va[ix]

sheet = wb["箱头表"]
for i in range(len(df_xiangtou)):
    va = df_xiangtou.iloc[i].values.T.tolist()
    for ix in range(len(va)):
        sheet[get_column_letter(ix + 1) + str(i + 2)] = va[ix]

sheet = wb["主机"]
for i in range(len(df_zhuji)):
    va = df_zhuji.iloc[i].values.T.tolist()
    for ix in range(len(va)):
        sheet[get_column_letter(ix + 1) + str(i + 2)] = va[ix]

sheet = wb["对重块"]
for i in range(len(df_duizhongkuai)):
    va = df_duizhongkuai.iloc[i].values.T.tolist()
    for ix in range(len(va)):
        sheet[get_column_letter(ix + 1) + str(i + 2)] = va[ix]

wb.save(r'E:\文件保存位置\工号基础信息.xlsx')

Save_Formats_xls(r'E:\文件保存位置\工号基础信息.xlsx', r'E:\文件保存位置\工号基础信息.xls')

# ifile.del_file(file=r'E:\文件保存位置\工号基础信息.xlsx')

# print(rpa_time.get_current_time())
