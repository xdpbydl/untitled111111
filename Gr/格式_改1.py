import openpyxl
import pandas as pd

# from win32com.client import Dispatch

excel_file = r'E:\TEMP\02GR\kefahuo\HANG\可发货通知书(代运）2020——导入2.0_1.xlsx'  # 格式文件
save_file = r'E:\TEMP\02GR\kefahuo\geshi\new_save.xlsx'  # 保存的新文件
data_fiel = r'E:\TEMP\02GR\kefahuo\geshi\0903.xlsx'  # 数据文件

# def just_open(filename):
#     xlApp = Dispatch("Excel.Application")
#     xlApp.Visible = False
#     xlBook = xlApp.Workbooks.Open(filename)
#     xlBook.Save()
#     xlBook.Close()


wb_yuan = openpyxl.load_workbook(excel_file, data_only=False)
# wb_data = openpyxl.load_workbook(data_fiel)
df_data = pd.read_excel(data_fiel, index=False, keep_default_na=False)
# print(wb_yuan.sheetnames)
sheet = wb_yuan["Sheet1"]
fille = openpyxl.styles.PatternFill('solid', fgColor='FFBB02')


# print (list(sheet.rows), list(sheet.columns))
def filegeshi(df_data, sheet):
    for i in range(len(df_data)):  # 列
        for r in range(len(df_data.columns)):  # 行
            # print(df_data.iloc[i, r], type(df_data.iloc[i, r]))
            # print(sheet.cell(row=r + 1, column=i + 2).value)
            # if x for x in ("")
            if openpyxl.utils.get_column_letter(r + 1) not in ["C", "D", "AG"]:  # 这三列存在公式避免写入
                sheet.cell(row=i + 2, column=r + 1, value=str(df_data.iloc[i, r]))
            if i + 1 == len(df_data):  # 最后一行填充颜色
                sheet.cell(row=i + 2, column=r + 1).fill = fille
    # print (sheet['d4'].value)
    # sheet = sheet[2:len(df_data)]
    sheet.delete_rows((len(df_data) + 2), 500)  # 删除模版，复制过来多余的行


filegeshi(df_data, sheet)
wb_yuan.save(save_file)

# just_open(save_file)

# sheet.cell(row=3, column=2).value
# for i in sheet[2]:
# print(i.value, end=" ")
# xl = openpyxl.Workbook()
# xls = xl.create_sheet(index=0)  # 新建一个excel，sheet表
# """通讯录"""
# list1 = ["年龄", "姓名", "手机号码"]  # 这里是我们需要的字段，可以根据需要自行添加，删除，修改
# for i in range(1, 40000):  # 这里是我们需要的行数。比如这里是4万
#     for j, value in enumerate(list1):
#         if j == 0:
#             xls.cell(i, 1).value = str(i)  # 根据特殊字段进行处理，输入不同的数值。可以自行添加其他需要的字段。
#         else:  # 其他情况或者没有特殊字段，直接使用添加列表中的值代替。
#             xls.cell(i, j + 1).value = value#
# xl.save(save_file)  # 保存我们生成的数据
