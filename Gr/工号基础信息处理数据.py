import pandas as pd
import openpyxl, time
import numpy as np

from openpyxl.utils import get_column_letter, column_index_from_string
from win32com.client import Dispatch

Excel_zhuangxiangInfo_Path = ''
Excel_xiangtou_Path = ''
Remote_file_path = ''
Excel_fahuodan_Path = ''
Excel_tongyongmoban = ''
Excel_tongyongmoban_xlsx = ''



# 另存为 因为Html网页格式表格不能读取，统一修改格式
def Save_Format_xls(filename):
    xlApp = Dispatch("Excel.Application")
    xlApp.Visible = False  # 后台运行
    xlApp.DisplayAlerts = False  # 不警报
    xlBook = xlApp.Workbooks.Open(filename)
    xlBook.SaveAs(filename, 56)  # 56   xls    51 xlsx
    xlBook.Close()


def Save_Formats_xls(filename, newfilename):
    xlApp = Dispatch("Excel.Application")
    xlApp.Visible = False  # 后台运行
    xlApp.DisplayAlerts = False  # 不警报
    xlBook = xlApp.Workbooks.Open(filename)
    xlBook.SaveAs(newfilename, 56)  # 56   xls    51 xlsx
    xlBook.Close()


start_t = time.time()

Save_Format_xls(Excel_zhuangxiangInfo_Path)

df1 = pd.read_excel(Excel_zhuangxiangInfo_Path)
# 删除不需要的列
df1.drop(["装箱单编号", "版本", "使用单位", "井道图号", "提升高度"], axis=1, inplace=True)

# ['工号', '订货单位', '型号', '层站门', '序号', '箱代码', '零件编号', '零件名称', '图表号', '作业单位','规格', '数量', '备注', '源图号作业', '物资编码', '参考图号']
# 帅选主机数据
df_zhuji = df1.loc[df1["箱代码"].str.strip() == "01-1"]

# 帅选对重块数据  and df1["零件名称"].str.strip() != "14#对重块箱"
df_duizhongkuai = df1.loc[(df1["箱代码"].str.strip() == "14-1") | (df1["箱代码"].str.strip() == "14-2")]
df_duizhongkuai = df_duizhongkuai.loc[df_duizhongkuai["零件名称"].str.strip() != "14#对重块箱"]

Save_Format_xls(Excel_xiangtou_Path)
df2 = pd.read_excel(Excel_xiangtou_Path, header=1)

# 获取层站门空白数据的索引
dfnaIndex = df2[pd.isna(df2["层站门"])].index

for i in range(len(dfnaIndex)):
    # 空白层站门 = 提升高度 + "MM "
    # df2["层站门"][dfnaIndex[i]] = str(int(df2["提升高度"][dfnaIndex[i]])) + "MM "
    df2.loc[dfnaIndex[i], "层站门"] = str(int(df2["提升高度"][dfnaIndex[i]])) + "MM "

# 箱头sheet表
df_xiangtou = df2[["工号", "订货单位", "型号", "层站门", "箱代码", "箱名", "备注"]]
# df_xiangtou["订货单位"] = df_xiangtou["订货单位"].astype("string")

# print(df_xiangtou.tail())

# [u"工号",u"订货单位",u"型号",u"层站门"]  借助pivot_table 透视表并且以count聚合 reset_index重置索引得到表结构
df_table = pd.pivot_table(df_xiangtou, index=["工号", "订货单位", "型号", "层站门"], aggfunc='count').reset_index()[
    ["工号", "订货单位", "型号", "层站门"]]
df_table["提货方式"] = "送货"
# print(df_table)


# 处理送货方式数据【工号信息】 df_table
sheetName = str(int(time.get_month())) + "-" + str(int(time.get_day()))

# 读取到共享文件目录下最新的发货计划
df_New = pd.DataFrame()
df_NewData = pd.read_excel(Remote_file_path, sheet_name=None, header=1)
for sheet_name, df in df_NewData.items():
    if sheetName == sheet_name:
        df_New = df[["工  号", "送货"]]

if not df_New.empty:
    # 删除Nan的行，因为阶梯的存在，会出现空行
    df_New = df_New.dropna(axis=0, how='all')
    for i in range(len(df_New)):
        if df_New["送货"][i] not in '送货':
            for ix in range(len(df_table)):
                if df_table.loc[ix, "工号"] == df_New["工  号"][i]:
                    df_table.loc[ix, "提货方式"] = df_New["送货"][i]

# 读取可发货通知单
df_fahuo = pd.read_excel(Excel_fahuodan_Path)
df_fahuo = df_fahuo[["电梯工号", "送货方式"]]
for i in range(len(df_fahuo)):
    if df_fahuo["送货方式"][i] not in '送货':
        for ix in range(len(df_table)):
            if df_table.loc[ix, "工号"] == df_fahuo["电梯工号"][i]:
                df_table.loc[ix, "提货方式"] = df_fahuo["送货方式"][i]

# va = []
wb = openpyxl.load_workbook(Excel_tongyongmoban, data_only=False)
sheet = wb["工号信息"]
for i in range(len(df_table)):
    va = df_table.iloc[i].values.T.tolist()
    for ix in range(len(va)):
        sheet[get_column_letter(ix + 1) + str(i + 2)] = va[ix]

sheet = wb["箱头表"]
for i in range(len(df_xiangtou)):
    va = df_xiangtou.iloc[i].values.T.tolist()
    for ix in range(len(va)):
        sheet[get_column_letter(ix + 1) + str(i + 2)] = va[ix]

sheet = wb["主机"]
for i in range(len(df_zhuji)):
    va = df_zhuji.iloc[i].values.T.tolist()
    for ix in range(len(va)):
        sheet[get_column_letter(ix + 1) + str(i + 2)] = va[ix]

sheet = wb["对重块"]
for i in range(len(df_duizhongkuai)):
    va = df_duizhongkuai.iloc[i].values.T.tolist()
    for ix in range(len(va)):
        sheet[get_column_letter(ix + 1) + str(i + 2)] = va[ix]

wb.save(Excel_tongyongmoban_xlsx)
Save_Formats_xls(Excel_tongyongmoban_xlsx, Excel_tongyongmoban_xlx)
# 删除工号基础信息xlsx
ifile.del_file(file=Excel_tongyongmoban_xlsx)

# 复制到共享目录文件下的基础信息
# 拼接文件保存路径，以日期来做保存
tvar0216223787237 = time.get_current_datetime_str(format='%Y%m%d')
Remote_PathInfo = Remote_Path + "\\" + str(tvar0216223787237)
# 检查路径是否存在
if not ifile.exist_file(filename=Remote_PathInfo):
    # 不存在时，创建文件夹
    ifile.create_dir(dir=Remote_PathInfo)

# 查找所有文件信息
file_ls = ifile.find_files(path=Excel_Path, include_file='*.xls', sub_dir=False, time_sort='Desc')

# for i in range(len(file_ls)):
# 获取文件地址需要转换格式  然后复制到共享目录上面
# ifile.copy_file(src_file=file_ls[i].replace('/','\\'),dst_file=Remote_PathInfo)
# 同时对文件进行删除
# ifile.del_file(file=file_ls[i].replace('/','\\'))

print(time.time() - start_t)