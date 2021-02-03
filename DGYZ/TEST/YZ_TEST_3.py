import openpyxl
import pandas as pd
from win32com.client import Dispatch

file_path = 'E:\\TEMP\\6TEST\\源文件\\'
save_path = 'E:\\TEMP\\6TEST\\'
model_path = 'E:\\TEMP\\6TEST\\model\\'

"""参数说明'type'为 'copy': r_row_len为总共需要取行数，r_col 为源数据取数的列名字符； 
'type'为 'openpyxl' 可以取公式代表的值。其中 r_row_len为总共需要取【开始行，结束行+1】，r_col【开始列，结束列+1】；
s_row, s_col都为保存的开始行列"""
excel_dict = {
    0: {'model_file': f'{model_path}各镇街统计工具202011.xlsx',
        'source_file': f'{file_path}A3701商业银行日报表_汇总表（人民币）_20201130_东莞市分行(44001444).xls',
        'save_file': f'{save_path}各镇街统计工具202011.xlsx',
        'type': 'copy', 'data': {'s_row': 1, 's_col': 1, 's_sheel': '1.A3701(人民币)',
                                 'r_header': None, 'r_row_len': 82, 'r_col': 'A:EA', 'r_sheel': 0}},
    1: {'model_file': f'{model_path}邮储银行各镇街存贷款统计表（2020-11-30）.xlsx',
        'source_file': f'{save_path}各镇街统计工具202011.xlsx',
        'save_file': f'{save_path}邮储银行各镇街存贷款统计表（2020-11-30）.xlsx',
        'type': 'openpyxl', 'data': {'s_row': 5, 's_col': 3, 's_sheel': '各镇街存贷款统计表',
                                     'r_header': 0, 'r_row_len': [2, 36], 'r_col': [24, 26], 'r_sheel': '结果'}},
    2: {'model_file': f'{save_path}各镇街统计工具202011.xlsx',
        'source_file': f'{file_path}A2411金融机构资产负债项目月报表（外币）_汇总表_20201130_东莞市分行(44001444).xls',
        'save_file': f'{save_path}各镇街统计工具202011.xlsx',
        'type': 'pd_data', 'data': {'s_row': 19, 's_col': 3, 's_sheel': '2.粘贴数据',
                                    'r_header': 0, 'r_row_len': 3, 'r_col': 'C:EA', 'r_sheel': ''}},
    3: {'model_file': f'{save_path}各镇街统计工具202011.xlsx',
        'source_file': f'{file_path}A3701商业银行日报表_汇总表（外币折人民币）_20201130_东莞市分行(44001444).xls',
        'save_file': f'{save_path}各镇街统计工具202011.xlsx',
        'type': 'pd_data', 'data': {'s_row': 26, 's_col': 3, 's_sheel': '2.粘贴数据',
                                    'r_header': 0, 'r_row_len': 3, 'r_col': 'C:EA', 'r_sheel': ''}},
}


#
# def Save_Format_xls(filename, type):
#     xlApp = Dispatch("Excel.Application")
#     xlApp.Visible = False  # 后台运行
#     xlApp.DisplayAlerts = False  # 不警报
#     xlBook = xlApp.Workbooks.Open(filename)
#     file_list = filename.split('.')
#     print(f'---123--{file_list}---')
#     if type == 'xls':
#         type_no = 56
#         file_list[-1] = 'xls'
#     elif type == 'xlsx':
#         type_no = 51
#         file_list[-1] = 'xlsx'
#     else:
#         print('转换文件类型不存在！')
#     xlBook.SaveAs(filename, type_no)  # 56   xls    51 xlsx
#     xlBook.Close()
#     print(f'---223--{file_list}---')
#     return '.'.join(file_list)


def source_data(source_file, header, cols, txt_list):
    """部分需要逻辑处理的取数"""
    df = pd.read_excel(source_file, header=header, keep_default_na=False)
    txt_list = txt_list
    df1 = df[df[cols].isin(txt_list)]
    df1 = df1.iloc[:, 2:]       #只取第3列及后面的数据
    return df1

def flag_no(i):
    """提供统一取数，方便函数调用"""
    if i ==2:
        df2 = source_data(source_file=excel_dict[2]['source_file'], header=2, cols='指标代码',
                      txt_list=['AA241122MG2', 'AA241122MG3', 'AA241122MI1'])
        return df2
    elif i ==3:
        df3 = source_data(source_file=excel_dict[3]['source_file'], header=2, cols='指标代码',
                          txt_list=['AA370137002', 'AA370136025', 'AA370136031'])
        return df3
    else:
        return 0


def just_open(filename):
    """防止读取excel公式为None"""
    xlApp = Dispatch("Excel.Application")
    xlApp.Visible = False
    xlBook = xlApp.Workbooks.Open(filename)
    xlBook.Save()
    xlBook.Close()


def r_s_excel(source_file, s_col, s_row, s_sheel, model_file, r_header, r_row_len, r_col, r_sheel, save_file, type, df):
    """从源文件 复制数据到 模板格式文件，再保存到结果文件"""
    print(f'--处理文件为：{source_file}-----')
    if type == 'openpyxl':
        just_open(source_file)
        source_excel = openpyxl.load_workbook(source_file, data_only=True)
        model_excel = openpyxl.load_workbook(model_file, data_only=False)
        source_sheet = source_excel[r_sheel]
        model_sheet = model_excel[s_sheel]
        dif_col = s_col - r_col[0]
        dif_row = s_row - r_row_len[0]
        for i in range(r_row_len[0], r_row_len[1]):  # 行
            for r in range(r_col[0], r_col[1]):  # 列
                print(source_sheet.cell(row=i, column=r).value)
                model_sheet.cell(row=i + dif_row, column=r + dif_col).value = source_sheet.cell(row=i, column=r).value

    elif type == 'copy':
        df = pd.read_excel(source_file, header=r_header, keep_default_na=False, sheet_name=r_sheel, usecols=r_col,
                           nrows=r_row_len - 1)
        model_excel = openpyxl.load_workbook(model_file, data_only=False)
        sheet = model_excel[s_sheel]
        for i in range(len(df)):  # 行
            for r in range(len(df.columns)):  # 列a
                if df.iloc[i, r] == '':
                    print(f'-----111---------{df.iloc[i, r]}---')
                    continue
                else:
                    print(f'-----222---------{df.iloc[i, r]}---')
                    sheet.cell(row=i + s_col, column=r + s_row, value=df.iloc[i, r])
    elif type == 'pd_data':
        df = df
        model_excel = openpyxl.load_workbook(model_file, data_only=False)
        sheet = model_excel[s_sheel]
        for i in range(len(df)):  # 行
            for r in range(len(df.columns)):  # 列a
                if df.iloc[i, r] == '':
                    print(f'-----111---------{df.iloc[i, r]}---')
                    continue
                else:
                    print(f'-----222---------{df.iloc[i, r]}---')
                    sheet.cell(row=i + s_col, column=r + s_row, value=df.iloc[i, r])
    model_excel.save(save_file)
    model_excel.close()

for i, v in excel_dict.items():
    # print(i, v['data'])
    source_file = v['source_file']
    model_file = v['model_file']
    save_file = v['save_file']
    data = v['data']

    r_s_excel(source_file, data['s_row'], data['s_col'], data['s_sheel'], model_file, data['r_header'],
              data['r_row_len'], data['r_col'], data['r_sheel'], save_file, v['type'], df=flag_no(i))
