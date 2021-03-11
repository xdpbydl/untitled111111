import openpyxl
from openpyxl.styles import PatternFill
import pandas as pd
from win32com.client import Dispatch

file_path_d1 = 'D:\\ZCXX\\3.1 DGYC\\1. 文档\整理文档\\1、监管月报一批(8份)\\'
save_path_d1 = 'E:\\TEMP\\6TEST\\DGYC1\\'
model_path_d1 = 'E:\\TEMP\\6TEST\\DGYC1\\model\\'
daily_name = '202006-监管报表.xlsx'
test = r'E:\TEMP\Desktop\111\test{aa}.xlsx'

file_path_d2 = 'E:\\TEMP\\6TEST\\DGYC2\\源文件\\'
save_path_d2 = 'E:\\TEMP\\6TEST\\DGYC2\\'
model_path_d2 = 'E:\\TEMP\\6TEST\\DGYC2\\model\\'

file_path_d4 = 'E:\\TEMP\\6TEST\\DGYC4\\源文件\\'
save_path_d4 = 'E:\\TEMP\\6TEST\\DGYC4\\'
model_path_d4 = 'E:\\TEMP\\6TEST\\DGYC4\\model\\'

"""
excel_dict参数说明'type'为 'pd_data': r_row_len为总共需要取行数，r_col 为源数据取数的列名字符； 
'type'为 'openpyxl' 可以取公式代表的值。其中 r_row_len为总共需要取【开始行，结束行+1】，r_col【开始列，结束列+1】；
s_row, s_col都为保存的开始行列
"""
excel_dict = {
    100: {'model_file': f'{model_path_d1}GF0103-191-存贷款明细报表-2020年12月月报（第一批次）.xlsx',
          'source_file': f'{file_path_d1}202006-监管报表.xlsx',
          'save_file': f'{save_path_d1}GF0103-191-存贷款明细报表-2020年12月月报（第一批次）.xlsx',
          'type': 'pd_data',
          'data': {'s_row': 6, 's_col': 3, 's_sheel': 'GF0103', 'r_header': 4, 'r_row_len': 32, 'r_col': 'C:E',
                   'r_sheel': '存贷明细'}},
    101: {'model_file': f'{model_path_d1}GF0100-181-资产负债项目统计表-2020年12月月报（第一批次）.xlsx',
          'source_file': f'{file_path_d1}202006-监管报表.xlsx',
          'save_file': f'{save_path_d1}GF0100-181-资产负债项目统计表-2020年12月月报（第一批次）.xlsx',
          'type': 'pd_data',
          'data': {'s_row': 5, 's_col': 3, 's_sheel': 'GF0100', 'r_header': 3, 'r_row_len': 130, 'r_col': 'C:E',
                   'r_sheel': 'G01'}},
    102: {'model_file': f'{model_path_d1}GF0102-201-贷款质量五级分类情况简表-2020年12月月报（第一批次）.xlsx',
          'source_file': f'{file_path_d1}202006-监管报表.xlsx',
          'save_file': f'{save_path_d1}GF0102-201-贷款质量五级分类情况简表-2020年12月月报（第一批次）.xlsx',
          'type': 'pd_data',
          'data': {'s_row': 6, 's_col': 3, 's_sheel': 'GF0102', 'r_header': 4, 'r_row_len': 11, 'r_col': 'C',
                   'r_sheel': '小五级'}},
    103: {'model_file': f'{model_path_d1}GF0109-161-存贷款月日均情况表-2020年12月月报（第一批次）.xlsx',
          'source_file': f'{file_path_d1}202006-监管报表.xlsx',
          'save_file': f'{save_path_d1}GF0109-161-存贷款月日均情况表-2020年12月月报（第一批次）.xlsx',
          'type': 'pd_data',
          'data': {'s_row': 6, 's_col': 3, 's_sheel': 'GF0109', 'r_header': 4, 'r_row_len': 11, 'r_col': 'C:E',
                   'r_sheel': '日均'}},
    104: {'model_file': f'{model_path_d1}SF6301-201-大中小微型企业贷款情况表-2020年12月月报（第一批次）.xlsx',
          'source_file': f'{file_path_d1}202006-监管报表.xlsx',
          'save_file': f'{save_path_d1}SF6301-201-大中小微型企业贷款情况表-2020年12月月报（第一批次）.xlsx',
          'type': 'pd_data',
          'data': {'s_row': 6, 's_col': 3, 's_sheel': 'SF6301', 'r_header': 4, 'r_row_len': 37, 'r_col': 'C:I',
                   'r_sheel': 'S6301'}},
    105: {'model_file': f'{model_path_d1}GF0101-161-表外业务情况表-2020年12月月报（第一批次）.xlsx',
          'source_file': f'{file_path_d1}202006-监管报表.xlsx',
          'save_file': f'{save_path_d1}GF0101-161-表外业务情况表-2020年12月月报（第一批次）.xlsx',
          'type': 'pd_data',
          'data': {'s_row': 6, 's_col': 4, 's_sheel': 'GF0101', 'r_header': 3, 'r_row_len': 39, 'r_col': 'D:E',
                   'r_sheel': '附注'}},
    200: {'model_file': f'{model_path_d2}各镇街统计工具202011.xlsx',
          'source_file': f'{file_path_d2}A3701商业银行日报表_汇总表（人民币）_20201130_东莞市分行(44001444).xls',
          'save_file': f'{save_path_d2}各镇街统计工具202011.xlsx',
          'type': 'pd_data', 'data': {'s_row': 4, 's_col': 3, 's_sheel': '1.A3701(人民币)',
                                      'r_header': 2, 'r_row_len': 82, 'r_col': 'C:EA', 'r_sheel': 0}},
    201: {'model_file': f'{model_path_d2}邮储银行各镇街存贷款统计表（2020-11-30）.xlsx',
          'source_file': f'{save_path_d2}各镇街统计工具202011.xlsx',
          'save_file': f'{save_path_d2}邮储银行各镇街存贷款统计表（2020-11-30）.xlsx',
          'type': 'openpyxl', 'data': {'s_row': 5, 's_col': 3, 's_sheel': '各镇街存贷款统计表',
                                       'r_header': 0, 'r_row_len': [2, 36], 'r_col': [24, 26], 'r_sheel': '结果'}},
    202: {'model_file': f'{save_path_d2}各镇街统计工具202011.xlsx',
          'source_file': f'{file_path_d2}A2411金融机构资产负债项目月报表（外币）_汇总表_20201130_东莞市分行(44001444).xls',
          'save_file': f'{save_path_d2}各镇街统计工具202011.xlsx',
          'type': 'pd_data', 'data': {'s_row': 19, 's_col': 3, 's_sheel': '2.粘贴数据',
                                      'r_header': 2, 'r_row_len': 3, 'r_col': 'C:EA', 'r_sheel': ''}},
    203: {'model_file': f'{save_path_d2}各镇街统计工具202011.xlsx',
          'source_file': f'{file_path_d2}A3701商业银行日报表_汇总表（外币折人民币）_20201130_东莞市分行(44001444).xls',
          'save_file': f'{save_path_d2}各镇街统计工具202011.xlsx',
          'type': 'pd_data', 'data': {'s_row': 26, 's_col': 3, 's_sheel': '2.粘贴数据',
                                      'r_header': 2, 'r_row_len': 3, 'r_col': 'C:EA', 'r_sheel': ''}},
    400: {'model_file': f'{model_path_d4}2020年11月金融统计信息（业务）.xlsx',  # 2020年11月金融统计信息（业务）.xlsx
          'source_file': f'{file_path_d4}存贷增减草稿2019.xlsx',
          'source_file_1': f'{file_path_d4}2020年11月金融统计信息（业务）.xlsx',
          'save_file': f'{save_path_d4}2020年11月金融统计信息（业务）.xlsx',
          'type': 'pd_data', 'data': {'s_row': 5, 's_col': 4, 's_sheel': '全市存贷增长',
                                      'r_header': 0, 'r_row_len': 43, 'r_col': 'C:F', 'r_sheel': '202011'}},
    401: {'model_file': f'{save_path_d4}2020年11月金融统计信息（业务）.xlsx',  # save_path_d4已经存在了部分数据，不使用model_file。
          'source_file': f'{file_path_d4}2020年11月重点业务报表-业务类zzj.xls',
          'save_file': f'{save_path_d4}2020年11月金融统计信息（业务）.xlsx',
          'type': 'pd_data', 'data': {'s_row': 6, 's_col': 5, 's_sheel': '储蓄余额按（管理支行和分局）',
                                      'r_header': 4, 'r_row_len': 41, 'r_col': 'E:R', 'r_sheel': '储蓄余额按（管理支行和分局）'}},
    402: {'model_file': f'{save_path_d4}2020年11月金融统计信息（业务）.xlsx',  # save_path_d4已经存在了部分数据，不使用model_file。
          'source_file': f'{file_path_d4}2020年11月重点业务报表-业务类zzj.xls',
          'save_file': f'{save_path_d4}2020年11月金融统计信息（业务）.xlsx',
          'type': 'pd_data', 'data': {'s_row': 5, 's_col': 7, 's_sheel': '储蓄余额按（网点）',
                                      'r_header': 3, 'r_row_len': 129, 'r_col': 'G:T', 'r_sheel': '储蓄余额按（网点）'}},
    403: {'model_file': f'{save_path_d4}2020年11月金融统计信息（业务）.xlsx',  # save_path_d4已经存在了部分数据，不使用model_file。
          'source_file': f'{file_path_d4}2020年11月重点业务报表-业务类zzj.xls',
          'save_file': f'{save_path_d4}2020年11月金融统计信息（业务）.xlsx',
          'type': 'pd_data', 'data': {'s_row': 5, 's_col': 3, 's_sheel': '余额分段简表',
                                      'r_header': 2, 'r_row_len': 7, 'r_col': 'C:L', 'r_sheel': '余额分段总'}},
    404: {'model_file': f'{save_path_d4}2020年11月金融统计信息（业务）.xlsx',  # save_path_d4已经存在了部分数据，不使用model_file。
          'source_file': f'{file_path_d4}2020年11月重点业务报表-业务类zzj.xls',
          'save_file': f'{save_path_d4}2020年11月金融统计信息（业务）.xlsx',
          'type': 'pd_data', 'data': {'s_row': 5, 's_col': 5, 's_sheel': '余额分段明细表',
                                      'r_header': 3, 'r_row_len': 132, 'r_col': 'E:V', 'r_sheel': '余额分段按网点'}},
    405: {'model_file': f'{save_path_d4}2020年11月金融统计信息（业务）.xlsx',
          'source_file': f'{file_path_d4}公司业务余额发展情况表-2020年11月.xls',
          'save_file': f'{save_path_d4}2020年11月金融统计信息（业务）.xlsx',
          'type': 'pd_data', 'data': {'s_row': 5, 's_col': 3, 's_sheel': '公司余额',
                                      'r_header': 3, 'r_row_len': 33, 'r_col': 'C:H', 'r_sheel': 0}},
    406: {'model_file': f'{save_path_d4}2020年11月金融统计信息（业务）.xlsx',  # 未完成！
          'source_file': f'{save_path_d2}各镇街统计工具202011.xlsx',
          'save_file': f'{save_path_d4}2020年11月金融统计信息（业务）_____.xlsx', 'is_sort': ['H', 'K', 'O', 'R'],
          'type': 'pd_data', 'data': {'s_row': 5, 's_col': 5, 's_sheel': '外币存款',
                                      'r_header': 0, 'r_row_len': 127, 'r_col': 'B,E,P,Q', 'r_sheel': '引用'}},
    407: {'model_file': f'{save_path_d4}2020年11月金融统计信息（业务）.xlsx',
          'source_file': f'{file_path_d4}2020年11月审查审批登记统计表（三农）.xls',
          'source_file_1': f'{file_path_d4}2020年11月金融统计信息（业务）.xlsx',
          'save_file': f'{save_path_d4}2020年11月金融统计信息（业务）.xlsx',
          'type': 'pd_data', 'data': {'s_row': 7, 's_col': 2, 's_sheel': '授信情况',
                                      'r_header': 2, 'r_row_len': 9, 'r_col': 'C:F,O:V,AE:AH', 'r_sheel': '汇总'}},
    408: {'model_file': f'{save_path_d4}2020年11月金融统计信息（业务）.xlsx',
          'source_file': f'{file_path_d4}11月票据统计（发宝平）支行.xlsx',
          'source_file_1': f'{file_path_d4}2020年11月金融统计信息（业务）.xlsx',  # 同比，取上月金额
          'source_file_2': f'{file_path_d4}20——年11月金融统计信息（业务）.xlsx',  # 环比，取去年相同月份的金额？
          'save_file': f'{save_path_d4}2020年11月金融统计信息（业务）.xlsx',
          'type': 'pd_data', 'data': {'s_row': 7, 's_col': 4, 's_sheel': '贴现业务',
                                      'r_header': 6, 'r_row_len': 10, 'r_col': 'C:G', 'r_sheel': 0}},
    409: {'model_file': f'{save_path_d4}2020年11月金融统计信息（业务）.xlsx',
          'source_file': f'{file_path_d4}2020年11月重点业务报表-业务类（标色）.xls',
          'save_file': f'{save_path_d4}2020年11月金融统计信息（业务）.xlsx', 'type': 'pd_data',
          'data': {'s_row': 6, 's_col': 5, 's_sheel': '代理业务', 'r_header': 4, 'r_row_len': 51, 'r_col': 'E:N', 'r_sheel': 0}},

    410: {'model_file': f'{save_path_d4}2020年11月金融统计信息（业务）.xlsx',
          'source_file': f'{file_path_d4}11月信用卡数据（陈宝平）.xls',
          'save_file': f'{save_path_d4}2020年11月金融统计信息（业务）.xlsx', 'type': 'pd_data',
          'data': {'s_row': 6, 's_col': 3, 's_sheel': '信用卡', 'r_header': 4, 'r_row_len': 52, 'r_col': 'C:S', 'r_sheel': 0}},

    411: {'model_file': f'{save_path_d4}2020年11月金融统计信息（业务）.xlsx',
          'source_file': f'{file_path_d4}2020年储汇代发工资情况表（汇总）.xls',
          'save_file': f'{save_path_d4}2020年11月金融统计信息（业务）.xlsx', 'type': 'pd_data',
          'data': {'s_row': 6, 's_col': 3, 's_sheel': '代发工资1', 'r_header': 2, 'r_row_len': 20, 'r_col': 'C:X', 'r_sheel': 0}},

    412: {'model_file': f'{save_path_d4}2020年11月金融统计信息（业务）.xlsx',
          'source_file': f'{file_path_d4}2020年11月重点业务报表-业务类.xls',
          'save_file': f'{save_path_d4}2020年11月金融统计信息（业务）.xlsx', 'type': 'pd_data', 'is_sort': ['R'],
          'data': {'s_row': 4, 's_col': 6, 's_sheel': '银信通', 'r_header': 1, 'r_row_len': 42, 'r_col': 'B,D,F,H,J,L,N,P,R,T,V,X,Z',
                   'r_sheel': '银信通1'}},

    413: {'model_file': f'{save_path_d4}2020年11月金融统计信息（业务）.xlsx',
          'source_file': f'{file_path_d4}2020年11月重点业务报表-业务类.xls',
          'save_file': f'{save_path_d4}2020年11月金融统计信息（业务）.xlsx', 'type': 'pd_data', 'is_sort': ['R'],
          'data': {'s_row': 5, 's_col': 6, 's_sheel': '手机银行', 'r_header': 1, 'r_row_len': 42, 'r_col': 'B, E, G, I, K, M, O, Q, S, U, W, Y, AA',
                   'r_sheel': '手机银行1'}},

    414: {'model_file': f'{save_path_d4}2020年11月金融统计信息（业务）.xlsx',
          'source_file': f'{file_path_d4}2020年11月重点业务报表-业务类.xls',
          'save_file': f'{save_path_d4}2020年11月金融统计信息（业务）.xlsx', 'type': 'pd_data', 'is_sort': ['R'],
          'data': {'s_row': 5, 's_col': 6, 's_sheel': '个人网银', 'r_header': 1, 'r_row_len': 42, 'r_col': 'B, E, G, I, K, M, O, Q, S, U, W, Y, AA',
                   'r_sheel': '个人网银1'}},

}


def source_data(source_file, header, cols, txt_list):
    """
    部分需要逻辑处理的取数
    """
    df = pd.read_excel(source_file, header=header, keep_default_na=False)
    txt_list = txt_list
    df1 = df[df[cols].isin(txt_list)]
    df1 = df1.iloc[:, 2:]  # 只取第3列及后面的数据
    return df1


def add_fixed_cols_rows(df, up_cols, add_col, fill_value, axis=1):
    """
    axis == 1增加列：在up_cols（字符串）列，后面增加固定的add_col（列list）列，,设置默认值为fill_value。注：列名重复会报错误
    axis == 0增加行：在up_cols（数字）行，后面增加固定的add_col（数字）个数行，设置默认值为fill_value
    """
    df = df.reset_index(drop=True)
    if axis == 1:
        df_columns = df.columns.tolist()
        for no, val in enumerate(add_col):
            df_columns.insert(df_columns.index(up_cols) + 1 + no, val)
        df = df.reindex(columns=df_columns, fill_value=fill_value)
    elif axis == 0:
        df_list = df.index.tolist()
        df_len = len(df_list)
        df['change'] = df_list
        df['change'] = df['change'].apply(lambda x: x + add_col if x >= up_cols else x)
        for i in range(add_col):
            df.loc[df_len + i] = [fill_value for _ in df.columns]
            df_list.append(df_len + i)
            df.loc[df_len + i, ['change']] = up_cols + i
        df = df.sort_values(by='change')
        df = df.set_index('change')
    df = df.reset_index(drop=True)
    return df


def flag_no(i):
    """
    提供统一取数，方便函数调用
    """
    source_file = excel_dict[i]['source_file']
    data = excel_dict[i]['data']
    if i in [201, 555]:  # openpyxl
        df = 0
    elif i == 202:
        df = source_data(source_file=source_file, header=data['r_header'], cols='指标代码',
                         txt_list=['AA241122MG2', 'AA241122MG3', 'AA241122MI1'])
    elif i == 203:
        df = source_data(source_file=source_file, header=data['r_header'], cols='指标代码',
                         txt_list=['AA370137002', 'AA370136025', 'AA370136031'])
    elif i == 200:
        df = pd.read_excel(excel_dict[i]['source_file'], header=data['r_header'], keep_default_na=False,
                           sheet_name=data['r_sheel'], usecols=data['r_col'], nrows=data['r_row_len'])
    elif i == 400:
        df_s = pd.read_excel(excel_dict[i]['source_file'], header=data['r_header'], keep_default_na=False,
                             sheet_name=data['r_sheel'], usecols=data['r_col'], nrows=data['r_row_len'], index_col=None)
        print(f'---------400----{df_s.columns}')
        df_s = df_s.reindex(columns=['本月各项存', '排名1', '本月个人', '排名2', '本月单位', '排名3', '本月各项贷款', '排名4'], fill_value='')
        df_s1 = pd.read_excel(excel_dict[i]['source_file_1'], header=3, keep_default_na=False,
                              sheet_name=data['s_sheel'], usecols='L:R', nrows=data['r_row_len'],
                              index_col=None)  # usecols='L:R'
        df_s1.rename(columns={'各项存款.1': '本年各项存', '排名.4': '排名5', '个人存款.1': '本年个人存款', '排名.5': '排名6',
                              '单位存款.1': '本年单位存款', '排名.6': '排名7', '各项贷款.1': '本年各项贷款'}, inplace=True)
        df_s1['排名5'] = ''
        df_s1['排名6'] = ''
        df_s1['排名7'] = ''
        df_s1["本年各项贷款"] = pd.to_numeric(df_s1["本年各项贷款"], errors='coerce')  # 字符转浮点类型
        df = df_s.join(df_s1)
        df['本年各项存'] = df['本月各项存'] + df['本年各项存']
        df['本年个人存款'] = df['本月个人'] + df['本年个人存款']
        df['本年单位存款'] = df['本月单位'] + df['本年单位存款']
        df['本年各项贷款'] = df['本月各项存'] + df['本年各项贷款']
    elif i == 405:
        df = pd.read_excel(excel_dict[i]['source_file'], header=3, keep_default_na=False,
                           sheet_name=data['r_sheel'], usecols=data['r_col'], nrows=data['r_row_len'])
        df = df.loc[0:data['r_row_len'] - 1]
        df.drop(index=[31], inplace=True)  # 删除第32行数据， ‘省行清分’

    elif i == 406:
        df_source = pd.read_excel(excel_dict[i]['source_file'], header=0, keep_default_na=False,
                                  sheet_name=data['r_sheel'], usecols=data['r_col'], nrows=data['r_row_len'])
        df_model = pd.read_excel(excel_dict[i]['model_file'], header=3, keep_default_na=False,
                                 sheet_name=data['s_sheel'], usecols='C,E,L', nrows=36, skiprows=[24])
        df_source["各项贷款"] = pd.to_numeric(df_source["各项贷款"], errors='coerce')  # 字符转浮点类型
        df_source["个人外币"] = pd.to_numeric(df_source["个人外币"], errors='coerce')  # 字符转浮点类型
        df = df_model.merge(df_source, on=['机构代码'])
        # df['个人存款'] = df['个人外币']
        # df['单位存款'] = df['单位外币']

        sort_dict = {'个人存款': '排名1', '单位存款': '排名2', '各项贷款': '排名3'}
        for key, val in sort_dict.items():
            df[val] = df[key].rank(method='first', numeric_only=True, ascending=False)
            df.loc[df[key] == 0, [val]] = ''  # 排名源为0，排名置空

        # df.drop(columns=['机构代码', '各项贷款', '单位外币', '个人外币'], inplace=True)

        df = df.reindex(columns=['个人存款', '上月余额', '本月新增', '排名1', '年初余额', '本年新增', '排名2', '单位存款', '单位上月余额', '本月增长', '排名3'], fill_value='')
        # df['排名1'] = df['个人存款'].apply(lambda x: x if x != '' or x != 0 else '')
        # df['排名2'] = df['单位存款'].apply(lambda x: x if x != '' or x != 0 else '')

        df = add_fixed_cols_rows(df, 20, 1, '', axis=0)
        # df = add_fixed_cols_rows(df, '本月新增', ['aa23', '本月232', '本月3333'], '', axis=1)

    elif i == 407:
        df_source = pd.read_excel(excel_dict[i]['source_file'], header=data['r_header'], keep_default_na=False,
                                  sheet_name=data['r_sheel'], usecols=data['r_col'], nrows=data['r_row_len'])
        col_name = ['审批信用笔数', '审批信用金额', '审批抵押笔数', '审批抵押金额', '审批小企业笔数', '审批小企业金额', '审批合计笔数', '审批合计金额', '通过信用笔数',
                    '通过信用金额', '通过抵押笔数', '通过抵押金额', '通过小企业笔数', '通过小企业金额', '通过合计笔数', '通过合计金额']
        df_source_a = pd.read_excel(excel_dict[i]['source_file_1'], header=19, keep_default_na=False,
                                    sheet_name=data['s_sheel'], usecols='B:Q', nrows=data['r_row_len'])
        df_source.columns = col_name
        df_source_a.columns = col_name
        df_sum = df_source + df_source_a
        df = pd.concat([df_source, df_sum])
        df = df.reset_index(drop=True)
        df = add_fixed_cols_rows(df, 9, 5, '', axis=0)

    elif i == 408:
        df_source = pd.read_excel(excel_dict[i]['source_file'], header=data['r_header'], keep_default_na=False,
                                  sheet_name=data['r_sheel'], usecols=data['r_col'], nrows=data['r_row_len'])
        df_source_a = pd.read_excel(excel_dict[i]['source_file_1'], header=data['r_header'] - 1, keep_default_na=False,
                                    sheet_name=data['s_sheel'], usecols='E,I,J,R', nrows=data['r_row_len'])
        df_col_name = list(df_source.columns)
        df_source = df_source.apply(pd.to_numeric, errors='coerce').fillna(0.0)  # 转换类型

        for lv in df_col_name:
            if '亿' in lv:
                df_source[lv] = df_source[lv] * 10000

        df_source.columns = ['库存张数', '库存金额', '贴现张数', '贴现金额', '收入']
        df_source_a.columns = ['上月贴现金额', '累计数量', '累计金额', '收入累计金额']

        col_name_all = ['贴现张数', '贴现金额', '本月排名', '上月', '环比', '本年贴现张数', '本年贴现金额', '本年排名', '去年', '同比', '库存张数', '库存金额', '收入', '本月收入排名', '本年收入']
        df_source = df_source.reindex(columns=col_name_all, fill_value='')
        df_source['上月'] = df_source_a['上月贴现金额']
        df_source['本年贴现张数'] = df_source['贴现张数'] + df_source_a['累计数量']
        df_source['本年贴现金额'] = df_source['贴现金额'] + df_source_a['累计金额']
        df_source['本年收入'] = df_source['收入'] + df_source_a['收入累计金额']
        df = df_source

    elif i == 411:
        df = pd.read_excel(excel_dict[i]['source_file'], header=data['r_header'], keep_default_na=False,
                           sheet_name=data['r_sheel'], usecols=data['r_col'], nrows=data['r_row_len'])
        df = df.loc[[0, 1, 2, 18, 19]]  # 只取这5行的值
        df.to_excel(test.format(aa=333))
        # input("#" * 18)

    elif i == 412:  # 缺，年初在网户数
        df_source = pd.read_excel(excel_dict[i]['source_file'], header=data['r_header'], keep_default_na=False,
                                  sheet_name=data['r_sheel'], usecols=data['r_col'], nrows=data['r_row_len'])
        df_model = pd.read_excel(excel_dict[i]['model_file'], header=2, keep_default_na=False,
                                 sheet_name=data['s_sheel'], usecols='E', nrows=51)
        df_source = df_source.fillna('')

        # 富华路，并入和大朗？
        df_source.iloc[df_source[df_source.机构号 == 441901019].index[0], 1:] += df_source.iloc[df_source[df_source.机构号 == 441901321].index[0], 1:]

        df_model = df_model.fillna('')
        df_model.columns = ['机构号']
        df_source['排名'] = ''
        df_source['年初在网户数'] = ''
        df_source.drop(index=data['r_row_len'] - 2, inplace=True)
        df_source.loc[data['r_row_len'] - 1, ['机构号']] = 1111  # 邮政小计，的'机构号' 为空，便于合平数据，要改名
        df = pd.merge(df_model, df_source, on='机构号', how='left')

        for val in df_source.columns:  # 取不为空的月份数据，来排名
            if df.loc[0, [val]].values == '':
                df['排名'] = df[mouth_col].rank(method='first', numeric_only=True, ascending=False)
                df.loc[df[mouth_col] == '', ['排名']] = ''  # 排名源为0，排名置空
                break
            mouth_col = val

        df.loc[50] = df_source.loc[41].tolist()  # 赋值，邮政小计 行数据
        df.drop(columns=['机构号'], inplace=True)
        df = df.fillna('')

    elif i == 413:  # 缺，年初在网户数
        df_source = pd.read_excel(excel_dict[i]['source_file'], header=data['r_header'], keep_default_na=False, sheet_name=data['r_sheel'],
                                  usecols=data['r_col'], nrows=data['r_row_len'])
        df_model = pd.read_excel(excel_dict[i]['model_file'], header=3, keep_default_na=False,
                                 sheet_name=data['s_sheel'], usecols='E', nrows=51)
        df_source = df_source.fillna('')

        # 富华路，并入和大朗？
        df_source.iloc[df_source[df_source.机构号 == 441901019].index[0], 1:] += df_source.iloc[df_source[df_source.机构号 == 441901321].index[0], 1:]

        df_model = df_model.fillna('')
        df_model.columns = ['机构号']
        df_source['排名'] = ''
        df_source['年初在网户数'] = ''
        df_source.drop(index=data['r_row_len'] - 2, inplace=True)
        df_source.loc[data['r_row_len'] - 1, ['机构号']] = 1111  # 邮政小计，的'机构号' 为空，便于合平数据，要改名
        df = pd.merge(df_model, df_source, on='机构号', how='left')

        for val in df_source.columns:  # 取不为空的月份数据，来排名
            if df.loc[0, [val]].values == '':
                df['排名'] = df[mouth_col].rank(method='first', numeric_only=True, ascending=False)
                df.loc[df[mouth_col] == '', ['排名']] = ''  # 排名源为0，排名置空
                break
            mouth_col = val

        df.loc[50] = df_source.loc[41].tolist()  # 赋值，邮政小计 行数据
        df.drop(columns=['机构号'], inplace=True)
        df = df.fillna('')

    elif i == 414:  # 缺，年初在网户数
        df_source = pd.read_excel(excel_dict[i]['source_file'], header=data['r_header'], keep_default_na=False, sheet_name=data['r_sheel'],
                                  usecols=data['r_col'], nrows=data['r_row_len'])
        df_model = pd.read_excel(excel_dict[i]['model_file'], header=3, keep_default_na=False,
                                 sheet_name=data['s_sheel'], usecols='E', nrows=51)
        df_source = df_source.fillna('')

        # 富华路，并入和大朗？
        df_source.iloc[df_source[df_source.机构号 == 441901019].index[0], 1:] += df_source.iloc[df_source[df_source.机构号 == 441901321].index[0], 1:]

        df_model = df_model.fillna('')
        df_model.columns = ['机构号']
        df_source['排名'] = ''
        df_source['年初在网户数'] = ''
        df_source.drop(index=data['r_row_len'] - 2, inplace=True)
        df_source.loc[data['r_row_len'] - 1, ['机构号']] = 1111  # 邮政小计，的'机构号' 为空，便于合平数据，要改名
        df = pd.merge(df_model, df_source, on='机构号', how='left')

        for val in df_source.columns:  # 取不为空的月份数据，来排名
            if df.loc[0, [val]].values == '':
                df['排名'] = df[mouth_col].rank(method='first', numeric_only=True, ascending=False)
                df.loc[df[mouth_col] == '', ['排名']] = ''  # 排名源为0，排名置空
                break
            mouth_col = val

        df.loc[50] = df_source.loc[41].tolist()  # 赋值，邮政小计 行数据
        df.drop(columns=['机构号'], inplace=True)
        df = df.fillna('')


    else:
        df = pd.read_excel(source_file, header=data['r_header'], keep_default_na=False, sheet_name=data['r_sheel'],
                           usecols=data['r_col'])
        df = df.loc[0:data['r_row_len'] - 1]
    return df


"""
    # df.to_excel(test.format(aa=333))
    # input("#" * 18)
"""


def just_open(filename):
    """
    防止读取excel公式为None
    """
    xlApp = Dispatch("Excel.Application")
    xlApp.Visible = False
    xlBook = xlApp.Workbooks.Open(filename)
    xlBook.Save()
    xlBook.Close()


def r_s_excel(dict_val, df):
    """
    从源文件 复制数据到 模板格式文件，再保存到结果文件
    (source_file, s_row, s_col, s_sheel, model_file, r_header, r_row_len, r_col, r_sheel, save_file, data_type, df)
    """
    source_file = dict_val['source_file']
    model_file = dict_val['model_file']
    save_file = dict_val['save_file']
    s_row = dict_val['data']['s_row']
    s_col = dict_val['data']['s_col']
    s_sheel = dict_val['data']['s_sheel']
    r_header = dict_val['data']['r_header']
    r_row_len = dict_val['data']['r_row_len']
    r_row_len = dict_val['data']['r_row_len']
    r_col = dict_val['data']['r_col']
    r_sheel = dict_val['data']['r_sheel']
    data_type = dict_val['type']

    print(f'--处理源文件文件为：{source_file}-----')
    print(f'--模板文件文件为：{model_file}-----')
    print(f'--保存文件为：{save_file}-----')
    if data_type == 'openpyxl':  # 从源文件为公式的单元格中取值
        just_open(source_file)
        source_excel = openpyxl.load_workbook(source_file, data_only=True)
        model_excel = openpyxl.load_workbook(model_file, data_only=False)
        source_sheet = source_excel[r_sheel]
        model_sheet = model_excel[s_sheel]
        dif_col = s_col - r_col[0]
        dif_row = s_row - r_row_len[0]
        for i in range(r_row_len[0], r_row_len[1]):  # 行
            for r in range(r_col[0], r_col[1]):  # 列
                print(source_sheet.cell(row=i, column=r).value)
                model_sheet.cell(row=i + dif_row, column=r + dif_col).value = source_sheet.cell(row=i, column=r).value
    elif data_type == 'pd_data':
        df = df
        model_excel = openpyxl.load_workbook(model_file, data_only=False)
        sheet = model_excel[s_sheel]
        for i in range(len(df)):  # 行
            for r in range(len(df.columns)):  # 列a
                if df.iloc[i, r] == '':
                    print(f'-----111---------{df.iloc[i, r]}---')
                    continue
                else:
                    try:
                        if openpyxl.utils.get_column_letter(r + s_col) in dict_val['is_sort'] and 0 < df.iloc[i, r] <= 5:
                            print(f'-----333---------{df.iloc[i, r]}---')
                            sheet.cell(row=i + s_row, column=r + s_col, value=df.iloc[i, r])
                            orange_fill = PatternFill(fill_type='solid', fgColor='FCE4D6')
                            sheet.cell(row=i + s_row, column=r + s_col).fill = orange_fill
                        else:
                            print(f'-----444---------{df.iloc[i, r]}---')
                            sheet.cell(row=i + s_row, column=r + s_col, value=df.iloc[i, r])
                    except:
                        print(f'-----222---------{df.iloc[i, r]}---')
                        sheet.cell(row=i + s_row, column=r + s_col, value=df.iloc[i, r])

    model_excel.save(save_file)
    model_excel.close()


for i, v in excel_dict.items():
    # # # if i < 100:
    # if i not in [400, 414]:
    #     continue
    print(f'--处理序号为：{i}-----')
    r_s_excel(v, df=flag_no(i))
