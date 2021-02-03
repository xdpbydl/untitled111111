import openpyxl
from win32com.client import Dispatch
import pandas as pd

file = 'E:\\TEMP\\6TEST\\源文件\\A3701商业银行日报表_汇总表（人民币）_20201130_东莞市分行(44001444).xls'
model_file = 'E:\\TEMP\\6TEST\\model\\邮储银行各镇街存贷款统计表（2020-11-30）.xlsx'
daily_file = 'D:\\ZCXX\\3.1 DGYC\\1. 文档\整理文档\\1、监管月报一批(8份)\\202006-监管报表.xlsx'
# source_file = 'E:\\TEMP\\6TEST\\各镇街统计工具202011.xlsx'
source_file = 'E:\\TEMP\\6TEST\\源文件\\A2411金融机构资产负债项目月报表（外币）_汇总表_20201130_东莞市分行(44001444).xls'
source_file_1 = 'E:\\TEMP\\6TEST\\源文件\\A3701商业银行日报表_汇总表（外币折人民币）_20201130_东莞市分行(44001444).xls'

save_file = 'E:\\TEMP\\6TEST\\111.xlsx'


def just_open(filename):
    """防止读取excel公式为None"""
    xlApp = Dispatch("Excel.Application")
    xlApp.Visible = False
    xlBook = xlApp.Workbooks.Open(filename)
    xlBook.Save()
    xlBook.Close()


def r_s_excel(source_file, s_col, s_row, s_sheel, model_file, r_header, r_row_len, r_col, r_sheel, save_file):
    just_open(source_file)
    source_excel = openpyxl.load_workbook(source_file, data_only=True)
    model_excel = openpyxl.load_workbook(model_file, data_only=False)
    source_sheet = source_excel[r_sheel]
    model_sheet = model_excel[s_sheel]
    dif_col = s_col -r_col[0]
    dif_row = s_row - r_row_len[0]
    for i in range(r_row_len[0], r_row_len[1]):  # 行
        for r in range(r_col[0], r_col[1]):  # 列
            print(source_sheet.cell(row=i, column=r).value)
            model_sheet.cell(row=i + dif_row, column=r + dif_col).value = source_sheet.cell(row=i, column=r).value

    model_excel.save(save_file)


# r_s_excel(source_file=source_file, s_col=3, s_row=5, s_sheel='各镇街存贷款统计表', model_file=model_file, r_header='', r_row_len=[2, 36], r_col=[24, 26], r_sheel='结果', save_file=save_file)
# r_s_excel(source_file=source_file, model_file=model_file, save_file=save_file)


def source_data(source_file, header, cols, txt_list):
    df = pd.read_excel(source_file, header=header, keep_default_na=False)
    txt_list = txt_list
    df1 = df[df[cols].isin(txt_list)]
    return df1

print(source_data(source_file=source_file, header=2, cols='指标代码', txt_list=['AA241122MG2', 'AA241122MG3', 'AA241122MI1']))
print(source_data(source_file=source_file_1, header=2, cols='指标代码', txt_list=['AA370137002', 'AA370136025', 'AA370136031']))
