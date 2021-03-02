import openpyxl

from openpyxl.styles import PatternFill

dir(openpyxl)
file = r'E:\TEMP\6TEST\test{}.xlsx'
print(file.format(1))
workbook = openpyxl.Workbook()
source_excel = openpyxl.load_workbook(file.format(1), data_only=False)
source_sheet = source_excel['Sheet1']

"""
方法一：使用fgColor填充指定单元格前景色
"""
sheet = workbook.create_sheet(index=0, title="Report")
sheet.append(['A', 'B', 'C'])
sheet.append(['D', 'E', 'F'])
orange_fill = PatternFill(fill_type='solid', fgColor='FFC125')
sheet.cell(row=1, column=1).fill = orange_fill
# workbook.save(file.format(3))
# workbook.close()


"""
方法二：使用start_color和end_color填充指定单元格颜色
注意：end_color可以省略，如果start_color和end_color的值不一致，则显示start_color指定颜色
"""
sheet.append(['A', 'B', 'C'])

sheet.append(['D', 'E', 'F'])

green_fill = PatternFill(start_color="AACF91", end_color="AACF91", fill_type="solid")

sheet.cell(row=1, column=2).fill = green_fill
# workbook.save(file.format(4))
# workbook.close()


def sort_set_color(colums={'个人金额': ['排名1', 5], '单位金额': ['排名2', 3]}):
    pass

source_sheet.cell(row=1, column=2).fill = green_fill
source_excel.save(file.format('_1'))
source_excel.close()