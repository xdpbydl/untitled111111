import pandas as pd
import numpy as np
import datetime
import openpyxl
from win32com.client import Dispatch

file_path = 'E:\\ZCXX\\东莞邮政\\1.需求\\给开发-\\'
jieju_file1 = f'{file_path}result2020-12-10----08-51-07(1).xls'
jieju_file = f'{file_path}temp\\result2020-12-10----08-51-07(1).xlsx'
renyuan_file = f'{file_path}全行人员花名册.xls'
model_file = f'{file_path}model\\小企业贷款日报模版.xls'
temp_file = f'{file_path}model\\temp.xls'
geishi_file = f'{file_path}model\\小企业贷款日报模版格式.xlsx'

last_file = f'{file_path}2020年12月8日小企业贷款日报.xls'  # 上一版本


def Save_Format_xls(filename1, filename):
    xlApp = Dispatch("Excel.Application")
    xlApp.Visible = False  # 后台运行
    xlApp.DisplayAlerts = False  # 不警报
    xlBook = xlApp.Workbooks.Open(filename1)
    xlBook.SaveAs(filename, 51)  # 56   xls    51 xlsx
    xlBook.Close()


for n in range(5):
    try:
        Save_Format_xls(jieju_file1, jieju_file)  # 转换时，小概率会报错
        print('***' * 8)
        break
    except:
        print(f'重新转换{n}次')

jieju_df = pd.read_excel(jieju_file, keep_default_na=False)

renyuan_df = pd.read_excel(renyuan_file, keep_default_na=False)
model_df = pd.read_excel(model_file, keep_default_na=False, header=2)
last_df = pd.read_excel(last_file, keep_default_na=False, header=2)
geshi_data = openpyxl.load_workbook(geishi_file, data_only=False)

last_df.columns = model_df.columns.tolist()  # 置换上一版本的列为模版的列名

last_df[['贷款结余笔数', '本日放款笔数', '贷款结余金额']] = last_df[['贷款结余笔数', '本日放款笔数', '贷款结余金额']].astype('float')
# last_df[['贷款结余金额']] = last_df[['贷款结余金额']].astype('float')

# #  日报不存在，但计算要用到的， 本日还款金额
last_df['本日还款金额'] = 0.00
model_df['本日还款金额'] = 0.00

model_name = model_df.columns.tolist()
model_name.insert(model_name.index('一级支行') + 1, '客户数量')  # 在 '一级支行' 列后面插入'客户数量'
model_df = model_df.reindex(columns=model_name, fill_value=0)
last_df = last_df.reindex(columns=model_name, fill_value=0)


def df_strip(df):
    df = df.copy()
    for c in df.columns:
        df = df.rename(columns={c: c.strip()})
        # 只去除以下列的空格
        if c in ['工作部门', '经办客户经理', '姓名', '借据状态', '客户编号', '一级支行', '风险分类']:
            if df[c].dtype == np.object:
                df[c] = pd.core.strings.str_strip(df[c])
    return df


# .借据余额 转为浮点类型
jieju_df["借据余额"] = jieju_df["借据余额"].str.replace(',', '')
jieju_df['借据余额'] = jieju_df['借据余额'].astype('float')

jieju_df = df_strip(jieju_df)
renyuan_df = df_strip(renyuan_df)

# 先替换'常平', '横沥'为同一地区
renyuan_df["工作部门"] = renyuan_df["工作部门"].replace(['常平', '横沥'], ['常平横沥', '常平横沥'])

# 部分人员对应不上，修改为城区的  '袁惠敏'
renyuan_set = renyuan_df.groupby('姓名')['姓名']
# jieju_set = jieju_df.groupby('经办客户经理')['经办客户经理']
a = []
for group_name, group_data in renyuan_set:
    a.append(group_name)
print(set(jieju_df[~jieju_df['经办客户经理'].isin(a)]['经办客户经理']))
jieju_df.loc[~jieju_df['经办客户经理'].isin(a), ['经办客户经理']] = '袁惠敏'

df1 = pd.merge(jieju_df, renyuan_df, how='left', left_on=["经办客户经理"], right_on=["姓名"], left_index=False,
               right_index=False)
# 客户数量
df1 = df1.set_index('借据起期')
df2 = df1.drop_duplicates(subset=["工作部门", "客户编号"], keep='first', inplace=False)
df3 = df2[df2.借据状态.isin(['正常', '逾期', '部分逾期'])]  # 只取这些有效数据
df3 = df3.groupby(["工作部门"])["客户编号"].count()
model_df = pd.merge(model_df, df3, how='left', left_on=["一级支行"], right_on=["工作部门"], left_index=False, right_index=False)
model_df.客户数量 = model_df.客户编号

cc = datetime.datetime.now()
month = cc.strftime("%Y/%m")
year = cc.strftime("%Y")
day = cc.strftime("%Y/%m/%d")
day = r'2020/12/9'

df1.借据金额 = df1.借据金额 / 10000
df1.借据余额 = df1.借据余额 / 10000
# fk 放款
fk_day = df1[day].groupby('工作部门')["借据金额"].agg({('本日放款笔数', 'count'), ('本日放款金额', 'sum')})
fk_month = df1[month].groupby('工作部门')["借据金额"].agg({('本月放款笔数', 'count'), ('本月放款金额', 'sum')})
fk_year = df1[year].groupby('工作部门')["借据金额"].agg({('本年放款笔数', 'count'), ('本年放款金额', 'sum')})

# jy_df 贷款结余 为计算 结余余额
jy_all = df1[df1.借据状态.isin(['正常', '逾期', '部分逾期'])].groupby('工作部门')["借据余额"].agg({('贷款结余笔数', 'count'), ('贷款结余金额', 'sum')})

# yq_df 逾期   取数—台账：借据状态—逾期、部分逾期     借据余额
yq_all = df1[df1.借据状态.isin(['逾期', '部分逾期'])].groupby('工作部门')["借据余额"].agg({('逾期金额', 'sum')})

# bl_df 不良  不良取数—台账：风险分类—次级、可疑、损失
bl_df = df1[df1.风险分类.isin(['次级', '可疑', '损失']) & (df1.借据状态.isin(['正常', '逾期', '部分逾期']))]  # 只取这些状态
bl_all = bl_df.groupby('工作部门')["借据余额"].agg({('不良金额', 'sum')})

# hc_df 合成临时表
hc_df = pd.concat([fk_day, fk_month, fk_year, jy_all, yq_all, bl_all, df3], axis=1)
col_list = hc_df.columns.tolist()

fk_join = pd.merge(model_df, hc_df, how='left', left_on=["一级支行"], right_index=True)
# col_list = ['本日放款笔数', '本日放款金额', '本月放款笔数', '本月放款金额', '本年放款笔数', '本年放款金额',
#             '贷款结余笔数', '贷款结余金额', '逾期金额', '不良金额']

for i in col_list:
    model_df[i] = fk_join[f'{i}_y']

model_df = model_df.fillna(0)
last_df = last_df.fillna(0)
# 本日还款=  上版本日贷款结余 + 本日放款 -  贷款结余
model_df.本日还款金额 = last_df.贷款结余金额 + model_df.本日放款金额 - model_df.贷款结余金额
# print(model_df.本日还款金额)
# print(last_df.贷款结余金额)
# print(model_df.本日放款金额)
# print(model_df.贷款结余金额); input()


# model_df.本月还款金额 = last_df.贷款结余金额 + model_df.本月放款金额 - model_df.贷款结余金额
# print(model_df.本日还款金额, last_df.贷款结余金额, model_df.本日放款笔数, model_df.贷款结余金额)

# 本日净增 = 本日放款 - 本日还款
model_df.本日净增金额 = model_df.本日放款金额 - model_df.本日还款金额

# print(model_df.本日净增金额)
# 本月净增 = 上版本本月净增  + 本日净增
model_df.本月净增金额 = last_df.本月净增金额 + model_df.本日净增金额
model_df.当季净增金额 = last_df.当季净增金额 + model_df.本日净增金额
model_df.本年净增金额 = last_df.本年净增金额 + model_df.本日净增金额

# 逾期金额比上月    逾期比上月 = 本日逾期 -（上版本逾期金额-上版本比上月）
model_df.比上月金额 = model_df.逾期金额 - (last_df.逾期金额 - last_df.比上月金额)
model_df.比去年金额 = model_df.逾期金额 - (last_df.逾期金额 - last_df.比去年金额)
# print(model_df)

# 增加部分合计列
col_list += ['客户数量', '本日净增金额', '本月净增金额', '当季净增金额', '本年净增金额', '比上月金额', '比去年金额']
# 计算总和
model_df.loc[9, col_list] = model_df.loc[(model_df['一级支行'] != '合计'), col_list].apply(lambda x: x.sum())

# 逾期率 = 逾期金额 / 贷款结余
model_df.逾期率 = model_df.逾期金额 / model_df.贷款结余金额
# 比上月逾期率  比上月 = 逾期率 - [（上版本逾期金额-上版本比上月) /（上版本贷款结余 - 上版本本月净增）]
model_df.比上月逾期率 = model_df.逾期率 - ((last_df.逾期金额 - last_df.比上月金额) / (last_df.贷款结余金额 - last_df.本月净增金额))
model_df.比去年逾期率 = model_df.逾期率 - ((last_df.逾期金额 - last_df.比去年金额) / (last_df.贷款结余金额 - last_df.本年净增金额))

# 不良率 同逾期率
model_df.比上月不良 = model_df.不良金额 - (last_df.不良金额 - last_df.比上月不良)
model_df.比去年不良 = model_df.不良金额 - (last_df.不良金额 - last_df.比去年不良)

model_df.不良率 = model_df.不良金额 / model_df.贷款结余金额
model_df.比上月不良率 = model_df.不良率 - ((last_df.不良金额 - last_df.比上月不良) / (last_df.贷款结余金额 - last_df.本月净增金额))
model_df.比去年不良率 = model_df.不良率 - ((last_df.不良金额 - last_df.比去年不良) / (last_df.贷款结余金额 - last_df.本年净增金额))

model_df.to_excel(temp_file, index=False)

save_day = cc.strftime("%Y{}%m{}%d{}").format('年', '月', '日')
excel_file1 = f'{file_path}model\\{save_day}小企业贷款日报.xls'

# 使用格式写入
sheet = geshi_data["Sheet1"]


def Date_Geshi(df_data, sheet):
    sheet.cell(1, 1).value = f'{save_day}各支行小企业贷款情况表'
    for i in range(10):  # 10行
        for r in range(40):  # 40列
            # 这几列、一行存在公式避免写入
            # if openpyxl.utils.get_column_letter(r + 1) not in ['G', 'J', 'L', 'N', 'P', 'R', "W", 'AA', 'AH', 'AO'] and i != 9:
            if sheet.cell(row=3, column=r + 1).value not in ['排名']:  # 第三行值不为 ‘排名’的列， 赋值
                sheet.cell(row=i + 4, column=r + 1, value=df_data.iloc[i, r])


Date_Geshi(model_df, sheet)

geshi_data.save(excel_file1)
