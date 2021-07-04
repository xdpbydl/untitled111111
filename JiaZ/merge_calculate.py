import pandas
import openpyxl
import os, time
from win32com.client import Dispatch

file_AB = "E:\\JiaZhi\\0.3 HUA\\AB.xlsx"
file_model = "E:\\JiaZhi\\0.3 HUA\\model\\model.xlsx"
file_path = "E:\\JiaZhi\\0.3 HUA\\量数据\\"

file_lists = [f"{file_path}{str(i)}.xls" for i in range(9)]


def Save_Format_xls(filename1):
    p, f = os.path.splitext(filename1);
    filename = f'{p}.xlsx'
    xlApp = Dispatch("Excel.Application")
    xlApp.Visible = False  # 后台运行
    xlApp.DisplayAlerts = False  # 不警报
    xlBook = xlApp.Workbooks.Open(filename1)
    xlBook.SaveAs(filename, 51)  # 56   xls    51 xlsx
    xlBook.Close()
    return filename


def del_files(path):
    # （使用 os.walk ,这个方法返回的是一个三元tupple(dirpath(string), dirnames(list), filenames(list)), 其中第一个为起始路径， 第二个为起始路径下的文件夹, 第三个是起始路径下的文件.）
    for root, dirs, files in os.walk(path):
        for name in files:
            if '.xls' in name and name != "8.xls":  # 判断某一字符串是否具有某一字串，可以使用in语句
                os.remove(os.path.join(root, name))  ## os.move语句为删除文件语句
                print(f'删除文件:,{os.path.join(root, name)}')
        if '8.xlsx' in files:
            print(files)
            try:
                os.rename(f'{root}\8.xls', f'{path}0.xls')  # 保留'8.xls'
            except:
                pass
            print(f'{root}\8.xlsx,改名为0.xlsx')


def kill_exe(exe_name):
    """
    杀死exe进程
    :param exe_name:进程名字
    :return:无
    """
    os.system('taskkill /f /t /im ' + exe_name)
    print("杀死进程{}".format(exe_name))


def merge(model_file, soc_file, new_fle, sheet=0, sheet_name='', data_only=False):
    try:
        soc_file = Save_Format_xls(soc_file)
    except:
        print(f"源文件{soc_file}不存在。")
    wb = openpyxl.load_workbook(soc_file, data_only=data_only)
    new_wb = openpyxl.load_workbook(model_file)
    soc_ws = wb.worksheets[sheet]

    #     print(soc_ws.title)
    #     soc_ws.title = sheet_name
    #     print(soc_ws.title)
    #     input()
    # print(soc_ws.title)
    # print(type(soc_ws.title))

    # new_ws = new_wb.create_sheet(soc_ws.title)

    if sheet_name != '':
        new_ws = new_wb[sheet_name]
    else:
        new_ws = new_wb[soc_ws.title]

    for x in range(1, soc_ws.max_row + 1):
        for y in range(1, soc_ws.max_column + 1):
            new_ws.cell(row=x, column=y, value=soc_ws.cell(row=x, column=y).value)
            # print(soc_ws.cell(row=x, column=y).value)
    new_wb.save(new_fle)
    new_wb.close()

    print(f"{soc_file},文件已经合并。")


is_first = 0  # 是否10:00点第一次执行,复制，删除excel.
kill_exe('wps.exe')  # 关闭excel or wps
kill_exe('excel.exe')  # 关闭excel or wps

if is_first:
    time.sleep(2)
    del_files(file_path)

merge(file_model, file_AB, file_AB, sheet=1, sheet_name='-1', data_only=True)
for n, i in enumerate(file_lists):
    merge(file_AB, i, file_AB)
