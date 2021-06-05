import openpyxl, os
from openpyxl.styles import Font, Alignment


def copy_ex(model_file, Soc_File, New_File):
    wb = openpyxl.load_workbook(Soc_File)
    new_wb = openpyxl.load_workbook(model_file)
    soc_ws = wb.worksheets[0]
    new_ws = new_wb.create_sheet(soc_ws.title)
    for x in range(1, 51):
        for y in range(1, 51):
            new_ws.cell(row=x, column=y, value=soc_ws.cell(row=x, column=y).value)
            # print(soc_ws.cell(row=x, column=y).value)
    new_wb.save(New_File)
    print("It is over")


Soc_File: str = "E:\\JiaZhi\\0.3 HUA\\量数据\\1.xlsx"
New_File = "E:\\JiaZhi\\0.3 HUA\\AB__1.xlsx"
Model_File = "E:\\JiaZhi\\0.3 HUA\\AB.xlsx"
# copy_ex(Model_File, Soc_File, New_File)

p, f = os.path.split(Soc_File);
print(" dir is: " + p)
print(" file is: " + f)

#  case 2:
drv, left = os.path.splitdrive(Soc_File);
print(" driver is: " + drv)
print(" left is: " + left)
#  case 3:
f, ext = os.path.splitext(Soc_File);
print(" f is: " + f)
print(" ext is: " + ext)
