from copy import copy
from openpyxl import load_workbook, Workbook

file_AB = "E:\\JiaZhi\\0.3 HUA\\AB__1.xlsx"
file_0 = "E:\\JiaZhi\\0.3 HUA\\量数据\\1.xlsx"

def replace_xls(src_file, tag_file, sheet_name):
    #        src_file是源xlsx文件，tag_file是目标xlsx文件，sheet_name是目标xlsx里的新sheet名称

    print("Start sheet %s copy from %s to %s" % (sheet_name, src_file, tag_file))
    wb = load_workbook(src_file)
    wb2 = load_workbook(tag_file)

    ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])
    # ws2 = wb2.create_sheet(sheet_name.decode('utf-8'))
    ws2 = wb2.create_sheet(sheet_name)

    max_row = ws.max_row  # 最大行数
    max_column = ws.max_column  # 最大列数

    wm = list(zip(ws.merged_cells))  # 开始处理合并单元格
    if len(wm) > 0:
        for i in range(0, len(wm)):
            cell2 = str(wm[i]).replace('(<MergeCell ', '').replace('>,)', '')
            print("MergeCell : %s" % cell2)
            ws2.merge_cells(cell2)

    for m in range(1, max_row + 1):
        ws2.row_dimensions[m].height = ws.row_dimensions[m].height
        for n in range(1, 1 + max_column):
            if n < 27:
                c = chr(n + 64).upper()  # ASCII字符,chr(65)='A'
            else:
                if n < 677:
                    c = chr(divmod(n, 26)[0] + 64) + chr(divmod(n, 26)[1] + 64)
                else:
                    c = chr(divmod(n, 676)[0] + 64) + chr(divmod(divmod(n, 676)[1], 26)[0] + 64) + chr(divmod(divmod(n, 676)[1], 26)[1] + 64)
            i = '%s%d' % (c, m)  # 单元格编号
            if m == 1:
                #				 print("Modify column %s width from %d to %d" % (n, ws2.column_dimensions[c].width ,ws.column_dimensions[c].width))
                ws2.column_dimensions[c].width = ws.column_dimensions[c].width
            try:
                getattr(ws.cell(row=m, column=c), "value")
                cell1 = ws[i]  # 获取data单元格数据
                ws2[i].value = cell1.value  # 赋值到ws2单元格
                if cell1.has_style:  # 拷贝格式
                    ws2[i].font = copy(cell1.font)
                    ws2[i].border = copy(cell1.border)
                    ws2[i].fill = copy(cell1.fill)
                    ws2[i].number_format = copy(cell1.number_format)
                    ws2[i].protection = copy(cell1.protection)
                    ws2[i].alignment = copy(cell1.alignment)
            except AttributeError as e:
                print("cell(%s) is %s" % (i, e))
                continue

    wb2.save(tag_file)

    wb2.close()
    wb.close()

replace_xls(file_0, file_AB, "1")